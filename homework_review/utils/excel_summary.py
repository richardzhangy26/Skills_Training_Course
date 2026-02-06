"""
Excel 评分表生成模块（统一 essay/thesis 与 exam_paper 类型）

根据 agent_code 自动选择数据提取策略：
- exam_paper: 使用 questionScores，full_mark 由各题 totalScore 累加
- essay_writing / thesis_writing: 使用 dimensionScores，full_mark 取 fullMark 字段
"""

from collections import defaultdict
from pathlib import Path
from typing import Optional


def extract_core_data(result: dict, agent_code: str = "") -> Optional[dict]:
    """从接口响应中提取核心评分数据（类型感知）"""
    if not isinstance(result, dict):
        return None

    report_data = result.get("data", result)
    if isinstance(report_data, dict) and "artifacts" in report_data:
        artifacts = report_data.get("artifacts") or []
        if artifacts:
            parts = artifacts[0].get("parts") or []
            core_data = parts[0].get("data", {}) if parts else {}
        else:
            core_data = {}
    else:
        core_data = report_data if isinstance(report_data, dict) else {}

    if not isinstance(core_data, dict):
        return None

    if agent_code == "exam_paper":
        question_scores = core_data.get("questionScores", [])
        full_mark = (
            sum(q.get("totalScore", 0) for q in question_scores)
            if question_scores
            else 100
        )
        return {
            "total_score": core_data.get("totalScore"),
            "full_mark": full_mark,
            "question_scores": question_scores,
        }
    else:
        return {
            "total_score": core_data.get("totalScore"),
            "full_mark": core_data.get("fullMark", 100),
            "dimension_scores": core_data.get("dimensionScores", []),
        }


def resolve_summary_output_root(file_paths, output_root: Optional[Path]) -> Path:
    if output_root:
        return output_root
    parents = {path.parent.resolve() for path in file_paths}
    if len(parents) == 1:
        return next(iter(parents)) / "review_results"
    return Path.cwd() / "review_results"


def generate_excel_summary(
    summary_items,
    file_paths,
    attempts: int,
    output_root: Optional[Path],
    agent_code: str = "",
):
    """生成 Excel 评分表（类型感知）

    Args:
        summary_items: 批改结果列表
        file_paths: 文件路径列表
        attempts: 每个文件的测评次数
        output_root: 输出根目录
        agent_code: 智能体类型代码 ("exam_paper" / "essay_writing" / "thesis_writing" 等)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("未找到 openpyxl，无法生成Excel评分表。请先安装: pip install openpyxl")
        return None

    output_root = resolve_summary_output_root(file_paths, output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    # 构建文件标签映射（处理重名）
    label_counts = defaultdict(int)
    label_by_path = {}
    for path in file_paths:
        key = str(path)
        if key not in label_by_path:
            base = path.stem
            label_counts[base] += 1
            label = base if label_counts[base] == 1 else f"{base}({label_counts[base]})"
            label_by_path[key] = label

    score_data = {}
    order_by_label = {}

    is_exam = agent_code == "exam_paper"

    for item in summary_items:
        if not item or not item.get("success"):
            continue
        core_data = extract_core_data(item.get("result", {}), agent_code)
        if not core_data:
            continue
        file_path = item.get("file_path", "")
        label = label_by_path.get(
            file_path, Path(file_path).stem if file_path else "unnamed"
        )

        if is_exam:
            entry = score_data.setdefault(
                label,
                {
                    "full_mark": core_data.get("full_mark", 100),
                    "total_scores": [None] * attempts,
                    "questions": {},
                },
            )
        else:
            entry = score_data.setdefault(
                label,
                {
                    "full_mark": core_data.get("full_mark", 100),
                    "total_scores": [None] * attempts,
                    "dimensions": {},
                },
            )

        order = order_by_label.setdefault(label, [])
        attempt_index = item.get("attempt_index", 0)

        if 1 <= attempt_index <= attempts:
            entry["total_scores"][attempt_index - 1] = core_data.get("total_score")

            if is_exam:
                for q in core_data.get("question_scores", []):
                    name = q.get("name") or f"题目{q.get('questionIndex', '?')}"
                    q_full = q.get("totalScore", 0)
                    if name not in order:
                        order.append(name)
                    q_entry = entry["questions"].setdefault(
                        name, {"full": q_full, "scores": [None] * attempts}
                    )
                    q_entry["scores"][attempt_index - 1] = q.get("score")
            else:
                for dim in core_data.get("dimension_scores", []):
                    name = dim.get("evaluationDimension") or "unnamed"
                    if name not in order:
                        order.append(name)
                    scores = entry["dimensions"].setdefault(name, [None] * attempts)
                    scores[attempt_index - 1] = dim.get("dimensionScore")

    if not score_data:
        print("未获取到可用于生成评分表的结果")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "评分表"

    col2_header = "评价项目" if is_exam else "评价维度"
    headers = ["档次/学生", col2_header] + [f"第{i}次" for i in range(1, attempts + 1)]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for path in file_paths:
        label = label_by_path.get(str(path))
        if label not in score_data:
            continue
        entry = score_data[label]
        order = order_by_label.get(label, [])

        full_mark = entry.get("full_mark", 100)
        try:
            full_mark_text = str(int(full_mark))
        except (TypeError, ValueError):
            full_mark_text = str(full_mark)
        total_label = f"总分（{full_mark_text}分）" if full_mark_text else "总分"

        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = total_label
        for idx, score in enumerate(entry.get("total_scores", []), start=3):
            ws.cell(row=row_idx, column=idx).value = score
        row_idx += 1

        if is_exam:
            for q_name in order:
                q_entry = entry["questions"].get(q_name)
                if not q_entry:
                    continue
                q_full = q_entry.get("full", 0)
                display_name = f"{q_name}（{q_full}分）"
                ws.cell(row=row_idx, column=1).value = ""
                ws.cell(row=row_idx, column=2).value = display_name
                scores = q_entry.get("scores", [None] * attempts)
                for idx, score in enumerate(scores, start=3):
                    ws.cell(row=row_idx, column=idx).value = score
                row_idx += 1
        else:
            for dim_name in order:
                ws.cell(row=row_idx, column=1).value = ""
                ws.cell(row=row_idx, column=2).value = dim_name
                scores = entry["dimensions"].get(dim_name, [None] * attempts)
                for idx, score in enumerate(scores, start=3):
                    ws.cell(row=row_idx, column=idx).value = score
                row_idx += 1

        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 18 if col_idx <= 2 else 12

    output_path = output_root / "评分表.xlsx"
    wb.save(output_path)
    print(f"评分表已生成: {output_path}")
    return output_path
